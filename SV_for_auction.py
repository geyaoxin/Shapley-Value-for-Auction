import itertools
import openpyxl
import numpy as np
import math
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
import copy

def can_reach(BI, PI):
    _BI=copy.deepcopy(BI)
    temp = set('S')
    CR = set()
    temp = temp.intersection(_BI)
    while len(temp) > 0:
        temp1 = set()
        while len(temp) > 0:
            temp2 = temp.pop()
            CR.add(temp2)
            _BI.discard(temp2)
            temp1 = temp1.union(PI[temp2][1])
        temp = temp1.intersection(_BI)
    return CR


def calculate_v_func(participantsInfo, allocateFunc, num_item=1):
    p_name = list(participantsInfo.keys())
    v_func = {}
    for size in range(1, len(p_name)+1):
        for coalition in itertools.combinations(p_name, size):
            v_func[coalition] = allocateFunc(
                participantsInfo, set(coalition), num_item)[0]
    return v_func


def first_price_allocate(participantsInfo, coalition, num_item=1):
    # now we just consider num_item=1
    # return: social welfare, (winner, payment)
    canReach = can_reach(coalition, participantsInfo)
    currentSW = 0
    winner = 'S'
    for j in canReach:
        if currentSW < participantsInfo[j][0][0]:
            winner = j
            currentSW = participantsInfo[j][0][0]
    return currentSW, (winner, currentSW)


def SW_of_permutation_first_price(participantsInfo, permutation, k=1):
    participantsName = list(participantsInfo.keys())
    currentSW = 0
    SW = []
    winner = []
    N = len(participantsName)-1

    for i in range(N+1):
        # analysis each participants in the randomPermutation order
        # the name of participants that we are considering at present
        beInvolved = set(permutation[0:i+1])
        # print(beInvolved)

        # judge which participants are valid
        # the name of participants that we can reach S at present
        canReach = can_reach(beInvolved, participantsInfo)
        # print(canReach)

        # get the current social welfare
        for j in canReach:
            currentSW = max(currentSW, participantsInfo[j][0][0])

        # record SW
        SW.append(currentSW)
    return SW


def SW_of_permutation_first_price_modified(participantsInfo, permutation, k=1):
    # ignore the last participant's valuation
    participantsName = list(participantsInfo.keys())
    currentSW = 0
    SW = []
    winner = []
    N = len(participantsName)-1

    for i in range(N+1):
        # analysis each participants in the randomPermutation order
        # the name of participants that we are considering at present
        beInvolved = set(permutation[0:i+1])
        # print(beInvolved)

        # judge which participants are valid
        # the name of participants that we can reach S at present
        canReach = can_reach(beInvolved, participantsInfo)
        # print(canReach)

        # get the current social welfare
        for j in canReach:
            if j != permutation[i]:
                currentSW = max(currentSW, participantsInfo[j][0][0])

        # record SW
        SW.append(currentSW)
    return SW


def calculate_marginal_SW(SW, permutation, participantsName):
    marginalSW_permutationOrder = []
    marginalSW_defaultOrder = [[]] * len(participantsName)
    for i in range(len(permutation)):
        if i == 0:
            marginal = SW[i]
        else:
            marginal = SW[i] - SW[i-1]
        marginalSW_permutationOrder.append(marginal)
        marginalSW_defaultOrder[participantsName.index(
            permutation[i])] = marginal
    return marginalSW_permutationOrder, marginalSW_defaultOrder

def random_mechanism(participantsInfo, permutation, k=1):
    participantsName = list(participantsInfo.keys())
    currentSW = 0
    SW = []
    winner = []

    N = len(participantsName)-1
    for i in range(N+1):
        # analysis each participants in the randomPermutation order
        # the name of participants that we are considering at present
        beInvolved = set(permutation[0:i+1])
        # print(beInvolved)

        # judge which participants are valid
        # the name of participants that we can reach S at present
        canReach = can_reach(beInvolved, participantsInfo)
        # print(canReach)

        # get the current social welfare
        for j in canReach:
            currentSW = max(currentSW, participantsInfo[j][0][0])
        # record SW
        SW.append(currentSW)
        # end condition
        if (currentSW <= participantsInfo[permutation[i]][0][0]) & (permutation[i] in canReach) & (permutation[i] != 'S'):
            winner = permutation[i]
            break
    # calculate marginal
    marginalSW_defaultOrder = [0] * len(participantsName)
    for i in range(len(SW)):
        if i == 0:
            marginal = SW[i]
        else:
            marginal = SW[i] - SW[i-1]
        marginalSW_defaultOrder[participantsName.index(
            permutation[i])] = marginal
    payment = [[0]]*len(participantsName)
    for i in range(len(participantsName)):
        payment[i] = 0 - marginalSW_defaultOrder[i]
    
    if winner != []:
        payment[participantsName.index(winner)] = SW[-1] + \
        payment[participantsName.index(winner)]

    
    # print results
    # print('Winner:', winner)
    # print('Default Order:',participantsName)
    # print('Marginal Social Welfare: ', marginalSW_defaultOrder)
    # print('Payment:', payment)
    return winner, marginalSW_defaultOrder, payment



def patched_random_mechanism(participantsInfo, permutation, k=1):
    participantsName = list(participantsInfo.keys())
    currentSW = 0
    SW = []
    winner = []
    utility= []
    N = len(participantsName)-1
    for i in range(N+1):
        # analysis each participants in the randomPermutation order
        # the name of participants that we are considering at present
        beInvolved = set(permutation[0:i+1])
        # print(beInvolved)

        # judge which participants are valid
        # the name of participants that we can reach S at present
        canReach = can_reach(beInvolved, participantsInfo)
        # print(canReach)

        # get the current social welfare
        for j in canReach:
            currentSW = max(currentSW, participantsInfo[j][0][0])
        # record SW
        SW.append(currentSW)
        # end condition
        if (currentSW <= participantsInfo[permutation[i]][0][0]) & (permutation[i] in canReach) & (permutation[i] != 'S'):
            winner = permutation[i]
            break
    # calculate marginal
    marginalSW_defaultOrder = [0] * len(participantsName)
    for i in range(len(SW)):
        if i == 0:
            marginal = SW[i]
        else:
            marginal = SW[i] - SW[i-1]
        marginalSW_defaultOrder[participantsName.index(
            permutation[i])] = marginal

    # if CANNOT stop
    if winner == []:
        highestBid = 0
        for participant in participantsName:
            if participantsInfo[participant][0][0] > highestBid:
                highestBid = participantsInfo[participant][0][0]
                highestBidder = participant
        newPermutation = np.array(permutation)
        newPermutation = np.delete(
            newPermutation, np.where(newPermutation == highestBidder))
        newPermutation = np.append(newPermutation, highestBidder)
        # print('CANNOT stop! New permutation:', newPermutation)
        currentSW = 0
        SW = []
        winner = []
        for i in range(N+1):
            # analysis each participants in the randomPermutation order
            # the name of participants that we are considering at present
            beInvolved = set(newPermutation[0:i+1])
            # print(beInvolved)

            # judge which participants are valid
            # the name of participants that we can reach S at present
            canReach = can_reach(beInvolved, participantsInfo)
            # print(canReach)

            # get the current social welfare
            for j in canReach:
                currentSW = max(currentSW, participantsInfo[j][0][0])
            # record SW
            SW.append(currentSW)
            # end condition
            if (currentSW <= participantsInfo[newPermutation[i]][0][0]) & (newPermutation[i] in canReach) & (newPermutation[i] != 'S'):
                winner = newPermutation[i]
                break
        # calculate marginal
        marginalSW_defaultOrder = [0] * len(participantsName)
        for i in range(len(SW)):
            if i == 0:
                marginal = SW[i]
            else:
                marginal = SW[i] - SW[i-1]
            marginalSW_defaultOrder[participantsName.index(
                newPermutation[i])] = marginal
    # calculate payment
    payment = [[0]]*len(participantsName)
    for i in range(len(participantsName)):
        payment[i] = 0 - marginalSW_defaultOrder[i]
    payment[participantsName.index(winner)] = SW[-1] + \
        payment[participantsName.index(winner)]

    # print results
    # print('Winner:', winner)
    # print('Default Order:',participantsName)
    # print('Marginal Social Welfare: ', marginalSW_defaultOrder)
    # print('Payment:', payment)
    return winner, marginalSW_defaultOrder, payment



def all_permutation_auction(participantsType, participantsInfo, Mechanism=patched_random_mechanism, output_to_xls=False, output_path="result.xlsx"):
    # the participantsType is the truthful information, and the participantsInfo is the reported type
    participantsName=list(participantsInfo.keys())
    N=len(participantsName)
    
    all_permutations=list(itertools.permutations(participantsName,N))
    result={}
    

    # S in the front
    # all_permutations=[tuple(["S"]+list(i)) for i in list(itertools.permutations(participantsName[1:],N-1))]
    # result={}

    for i in all_permutations:
        result[i]=Mechanism(participantsInfo, i)
    
    utility={}
    for i in all_permutations:
        winner=result[i][0]
        u=[-j for j in result[i][2]] 
        if winner != []:
            winner_index=participantsName.index(winner)
            item_value=participantsType[winner][0][0]
            u[winner_index]=u[winner_index]+item_value   
        else:
            for j in range(len(u)):
                u[0]=u[0]-u[j]            
        utility[i]=u
    
    return result,utility


def dictory_average(dictory):
    
    avg_dict=[0 for i in list(dictory.keys())[0]]
    for i in dictory:
        for j in range(len(avg_dict)):
            avg_dict[j]=avg_dict[j]+dictory[i][j]
    for i in range(len(avg_dict)):
        avg_dict[i]=avg_dict[i]/len(dictory.keys())
    return avg_dict        



def misreport_price(participantsType,agent,price_range, Mechanism=patched_random_mechanism, show_plot=False):
    #the agent misreports the price in price_range, show the utility of the agent 
    #price range is a list of price the agent misreport
    data={}
    participantsName=list(participantsType.keys())
    agent_index=participantsName.index(agent)
    for price in price_range:
        # participantsInfo=participantsType.copy()
        participantsInfo=copy.deepcopy(participantsType)
        participantsInfo[agent][0][0]=price
        result,utility=all_permutation_auction(participantsType,participantsInfo,Mechanism)
        avg_utility=dictory_average(utility)
        data[price]=avg_utility

    if show_plot==True:
        x_data=[i for i in data]
        y_data=[data[i][agent_index] for i in data]
        plt.plot(x_data,y_data,marker='o')
        plt.show()
    return data



def misreport_neighbors(participantsType,agent, Mechanism=patched_random_mechanism):
    data={}
    neighbors=copy.deepcopy(participantsType[agent][1])
    possible_reports=[]
    for i in range(len(neighbors)+1):
        for j in list(itertools.combinations(neighbors,i)):
            possible_reports.append(j)
    for i in possible_reports:
        participantsInfo=copy.deepcopy(participantsType)
        participantsInfo[agent]=(participantsInfo[agent][0],copy.deepcopy(i))
        result,utility=all_permutation_auction(participantsType,participantsInfo,Mechanism)
        avg_utility=dictory_average(utility)
        data[i]=avg_utility
    return data





# def compute_SV_for_auction(participantsInfo,participantsReport, Mechanism=patched_random_mechanism, output_to_xls=True, output_path="result.xlsx"):
#     participantsName = participantsInfo.keys()
#     all_permutations = itertools.permutations(
#         participantsName, len(participantsName))
#     avg = ([0 for j in range(len(participantsName))],
#            [0 for j in range(len(participantsName))])

    
#     if output_to_xls:
#         wb = openpyxl.Workbook()

#         wb.create_sheet(index=0, title="1")
#         sheet = wb.worksheets[0]

#         count = 1
#         n = 1
#         sheet.cell(count, n).value = "winner"
#         n = n+1

#         sheet.cell(count, n).value = "Marginal"
#         n = n+1
#         for i in participantsName:
#             sheet.cell(count, n).value = i
#             n = n+1

#         sheet.cell(count, n).value = "Payment"
#         n = n+1
#         for i in participantsName:
#             sheet.cell(count, n).value = i
#             n = n+1

#         sheet.cell(count, n).value = "utility"
#         n = n+1
#         for i in participantsName:
#             sheet.cell(count, n).value = i
#             n = n+1


#     for i in all_permutations:
#         result = patched_random_mechanism(participantsReport, i)
#         for j in range(len(participantsName)):
#             avg[0][j] = avg[0][j]+result[1][j]
#             avg[1][j] = avg[1][j]+result[2][j]
#         if output_to_xls:
#             n=1
#             for j in result:
#                 for k in range(1,len(j)+1):
#                     sheet.cell(count,n).value=j[k-1]
#                     n=n+1
#                 n=n+1
#         count=count+1
#     for j in range(len(participantsName)):
#         avg[0][j]=avg[0][j]/(count-2)
#         avg[1][j]=avg[1][j]/(count-2)

#     if output_to_xls:
#         wb.save(output_path)
    
#     return avg
